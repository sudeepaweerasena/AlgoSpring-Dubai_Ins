import openpyxl
import pandas as pd
from datetime import datetime

# File paths and sheet names
OUTPUT_FILE_PATH = "C:\\Users\\sudeepa.w\\Documents\\GitHub\\AlgoSpring-Dubai_Ins\\MemberUpload - Dubaicare.xlsx"
OUTPUT_SHEET_NAME = "Sheet1"
INPUT_FILE_PATH = "C:\\Users\\sudeepa.w\\Documents\\GitHub\\AlgoSpring-Dubai_Ins\\CensusData.xlsx"
INPUT_SHEET_NAME = "Sheet1"

def nlg_transfer_medical_data():
    try:
        input_df = pd.read_excel(INPUT_FILE_PATH, sheet_name=INPUT_SHEET_NAME)
        print("Initial Input Data:")
        print(input_df.head())

    except FileNotFoundError:
        print(f"Error: The file {INPUT_FILE_PATH} does not exist.")
        return
    except ValueError:
        print(f"Error: The sheet {INPUT_SHEET_NAME} does not exist in {INPUT_FILE_PATH}.")
        return

    # Replace 'Principal' with 'Employee' in the 'Relation' column
    input_df['Relation'] = input_df['Relation'].replace('Principal', 'Employee')
    print("After Replacing 'Principal' with 'Employee':")
    print(input_df.head())


    # Convert Salary Type to 'Yes' or 'No'
    input_df['Visa Issued Emirates'] = input_df['Visa Issued Emirates'].apply(lambda visa_location: 'DXB' if visa_location == 'Dubai' else ('No' if visa_location == 'AhuDubai' else visa_location))
    print("After Converting Salary Type:")
    print(input_df.head())

    # Rename columns as per template needs
    column_mapping = {
        "Relation": "Relation",
        "Gender": "Gender",
        "DOB": "DOB",
        "Category": "Category",
        "Marital status": "Marital Status",
        "Visa Issued Emirates": "Visa Location",
        "Salary Type": "Salary Type"
    }
    input_df.rename(columns=column_mapping, inplace=True)

    # Load the output workbook and sheet
    try:
        output_wb = openpyxl.load_workbook(OUTPUT_FILE_PATH)
        output_ws = output_wb[OUTPUT_SHEET_NAME]
    except FileNotFoundError:
        print(f"Error: The file {OUTPUT_FILE_PATH} does not exist.")
        return
    except KeyError:
        print(f"Error: The sheet {OUTPUT_SHEET_NAME} does not exist in {OUTPUT_FILE_PATH}.")
        return

    # Apply the DOB formatting
    for index, row in input_df.iterrows():
        dob = row['DOB']
        if pd.notnull(dob):
            dob_converted = pd.to_datetime(dob, errors='coerce')
            if pd.notnull(dob_converted):
                formatted_dob = dob_converted.strftime("%d-%b-%y")  # Format changed to d-MMM-yy

            else:
                formatted_dob = "Invalid DOB"
        else:
            formatted_dob = "Invalid DOB"

        output_ws.cell(row=index + 2, column=3).value = formatted_dob
        output_ws.cell(row=index + 2, column=1).value = index + 1
        output_ws.cell(row=index + 2, column=7).value = row["Marital Status"]
        output_ws.cell(row=index + 2, column=2).value = row["Gender"]
        # output_ws.cell(row=index + 2, column=7).value = row["Relation"]
        output_ws.cell(row=index + 2, column=4).value = "Enhanced"
        output_ws.cell(row=index + 2, column=5).value = row["Visa Location"]
        category_letter = row["Category"]
        category_mapping = {'A': "Category A", 'B': "Category B", 'C': "Category C"}
        category_value = category_mapping.get(category_letter, "Unknown")
        output_ws.cell(row=index + 2, column=6).value = category_value

    output_wb.save(OUTPUT_FILE_PATH)
    print(f"Data successfully written to {OUTPUT_FILE_PATH}")

# # Call the function to execute
# nlg_transfer_medical_data()
